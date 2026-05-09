"""Generate sample .xlsx files for testing the xlaudit dashboard."""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from pathlib import Path

out = Path(__file__).parent / "samples"
out.mkdir(exist_ok=True)

# ── File 1: budget_2024.xlsx (complex) ──────────────────────────────────
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Summary"
ws1["A1"] = "Budget 2024"
for r in range(2, 40):
    ws1.cell(r, 1, f"=SUM(B{r}:D{r})")
    ws1.cell(r, 2, r * 100)
    ws1.cell(r, 3, f"=Revenue!B{r}*0.8")
    ws1.cell(r, 4, f"=NOW()")

ws2 = wb.create_sheet("Revenue")
for r in range(1, 50):
    ws2.cell(r, 1, f"Item {r}")
    ws2.cell(r, 2, r * 250)
    ws2.cell(r, 3, f"=Summary!A{r}")
    ws2.cell(r, 4, f"=TODAY()")

ws3 = wb.create_sheet("Costs")
for r in range(1, 30):
    ws3.cell(r, 1, f"=Revenue!B{r}*0.3")
    ws3.cell(r, 2, f"=Summary!A{r}")

ws4 = wb.create_sheet("Forecast")
for r in range(1, 20):
    ws4.cell(r, 1, f"=RAND()")
    ws4.cell(r, 2, f"=Revenue!B{r}+Costs!A{r}")
    ws4.cell(r, 3, f"=[External_Model.xlsx]Assumptions!C{r}")

ws5 = wb.create_sheet("Validation")
ws6 = wb.create_sheet("Archive")

for name, ref in [("TotalRevenue", "Revenue!$B$1:$B$50"), ("TotalCosts", "Costs!$A$1:$A$30"),
                   ("ForecastRange", "Forecast!$A$1:$C$20"), ("BudgetYear", "Summary!$A$1")]:
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)

wb.save(str(out / "budget_2024.xlsx"))
print("Created budget_2024.xlsx")

# ── File 2: sales_summary.xlsx (simple) ─────────────────────────────────
wb2 = openpyxl.Workbook()
ws = wb2.active
ws.title = "Q4 Sales"
for r in range(1, 20):
    ws.cell(r, 1, f"Product {r}")
    ws.cell(r, 2, r * 50)
    ws.cell(r, 3, f"=B{r}*1.1")

ws2b = wb2.create_sheet("Totals")
for r in range(1, 5):
    ws2b.cell(r, 1, f"='Q4 Sales'!B{r}")

dn = DefinedName("SalesRange", attr_text="'Q4 Sales'!$B$1:$B$20")
wb2.defined_names.add(dn)

wb2.save(str(out / "sales_summary.xlsx"))
print("Created sales_summary.xlsx")

# ── File 3: kpi_dashboard.xlsx (medium-high) ────────────────────────────
wb3 = openpyxl.Workbook()
ws1 = wb3.active
ws1.title = "KPIs"
for r in range(1, 60):
    ws1.cell(r, 1, f"=Targets!A{r}")
    ws1.cell(r, 2, f"=Actuals!B{r}")
    ws1.cell(r, 3, f'=IF(B{r}>A{r},"OK","MISS")')
    ws1.cell(r, 4, f'=INDIRECT("Targets!C"&ROW())')

ws2 = wb3.create_sheet("Targets")
for r in range(1, 60):
    ws2.cell(r, 1, r * 1000)
    ws2.cell(r, 2, f"=KPIs!B{r}")
    ws2.cell(r, 3, f"Zone {(r % 5) + 1}")

ws3 = wb3.create_sheet("Actuals")
for r in range(1, 60):
    ws3.cell(r, 1, f"=OFFSET(Targets!A{r},0,1)")
    ws3.cell(r, 2, r * 980)

ws4 = wb3.create_sheet("Trends")
for r in range(1, 30):
    ws4.cell(r, 1, f"=KPIs!B{r}-KPIs!A{r}")
    ws4.cell(r, 2, f"=[Market_Data.xlsx]Prices!D{r}")
    ws4.cell(r, 3, f"=NOW()")

for name, ref in [("KPI_Range", "KPIs!$A$1:$D$60"), ("TargetZones", "Targets!$C$1:$C$60"),
                   ("TrendData", "Trends!$A$1:$B$30")]:
    dn = DefinedName(name, attr_text=ref)
    wb3.defined_names.add(dn)

wb3.save(str(out / "kpi_dashboard.xlsx"))
print("Created kpi_dashboard.xlsx")

print(f"\nAll samples written to: {out.resolve()}")
