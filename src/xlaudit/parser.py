"""Low-level workbook parsing with openpyxl."""

from __future__ import annotations

import re
from typing import List, Set, Tuple

import openpyxl
from openpyxl.workbook import Workbook

# ── Regex patterns ──────────────────────────────────────────────────────────
# Matches the bracketed filename in external references like [Budget.xlsx]
_RE_EXTERNAL_LINK = re.compile(
    r"""\[           # opening bracket
        ([^\]]+)     # capture: filename or path
    \]               # closing bracket
    """,
    re.VERBOSE,
)

# Matches the *full* external reference including trailing sheet!cell,
# used to strip them before counting cross-sheet refs.
_RE_EXTERNAL_FULL = re.compile(
    r"""'?\[           # optional quote then opening bracket
        [^\]]+         # filename or path
    \]                 # closing bracket
    [^!]*              # sheet name (anything up to !)
    ![^,\s\)]*         # !cell-reference
    '?                 # optional closing quote
    """,
    re.VERBOSE,
)

# Matches cross-sheet references like Sheet1!A1 or 'My Sheet'!B2
# but NOT external links (which start with [)
_RE_CROSS_SHEET = re.compile(
    r"""(?<!\[)       # not preceded by [
    (?:'[^']+' | [A-Za-z_]\w*)   # sheet name (quoted or unquoted)
    !                # exclamation mark
    (?:\$?[A-Z]{1,3}\$?\d+)  # cell reference
    """,
    re.VERBOSE,
)

# Volatile functions (case-insensitive)
_VOLATILE_FUNCTIONS = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "INDIRECT", "OFFSET"}
_RE_VOLATILE = re.compile(
    r"\b(" + "|".join(_VOLATILE_FUNCTIONS) + r")\s*\(",
    re.IGNORECASE,
)


def load_workbook(path: str) -> Workbook:
    """Open an .xlsx workbook in read-only / data-only=False mode."""
    return openpyxl.load_workbook(path, read_only=True, data_only=False)


def iter_formulas(ws) -> List[str]:
    """Yield all formula strings in a worksheet."""
    formulas: List[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.value)
    return formulas


def detect_external_links(formulas: List[str]) -> Set[str]:
    """Return unique external file references found across *formulas*."""
    links: Set[str] = set()
    for f in formulas:
        for m in _RE_EXTERNAL_LINK.finditer(f):
            links.add(m.group(1))
    return links


def count_volatile(formulas: List[str]) -> int:
    """Count total volatile‑function invocations across *formulas*."""
    count = 0
    for f in formulas:
        count += len(_RE_VOLATILE.findall(f))
    return count


def count_cross_sheet_refs(formulas: List[str]) -> int:
    """Count cross-sheet cell references (excluding external links)."""
    count = 0
    for f in formulas:
        # Remove external-link portions first so we don't double-count
        cleaned = _RE_EXTERNAL_FULL.sub("", f)
        count += len(_RE_CROSS_SHEET.findall(cleaned))
    return count


# Captures the sheet name from cross-sheet references
_RE_CROSS_SHEET_TARGET = re.compile(
    r"""(?<!\[)
    (?:'([^']+)' | ([A-Za-z_]\w*))   # group 1: quoted, group 2: unquoted
    !
    \$?[A-Z]{1,3}\$?\d+
    """,
    re.VERBOSE,
)


def detect_cross_sheet_targets(formulas: List[str]) -> dict[str, int]:
    """Return ``{target_sheet_name: ref_count}`` for cross-sheet refs.

    External-link references are excluded.
    """
    targets: dict[str, int] = {}
    for f in formulas:
        cleaned = _RE_EXTERNAL_FULL.sub("", f)
        for m in _RE_CROSS_SHEET_TARGET.finditer(cleaned):
            name = m.group(1) or m.group(2)
            if name:
                targets[name] = targets.get(name, 0) + 1
    return targets


def get_named_ranges(wb: Workbook) -> List[str]:
    """Return list of defined-name labels in the workbook."""
    try:
        return [dn.name for dn in wb.defined_names.definedName]
    except AttributeError:
        return list(wb.defined_names.keys()) if hasattr(wb.defined_names, "keys") else []

